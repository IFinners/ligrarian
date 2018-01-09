#!/usr/bin/env python3

"""Automatically updates Books Read spreadsheet and Goodreads account."""

import configparser
import sys
import time
import datetime
from datetime import timedelta
import requests
import openpyxl
import bs4
from selenium import webdriver
from selenium.webdriver.common.keys import Keys


def get_date():
    """Return the date the book was read formatted (DD/MM/YY)."""
    today = datetime.datetime.now()

    date_day = today.day
    if len(str(date_day)) == 1:
        date_day = '0' + str(date_day)
    date_month = today.month
    if len(str(date_month)) == 1:
        date_month = '0' + str(date_month)

    date_year = str(today.year)[-2:]

    if len(sys.argv) == 4:
        date = '{}/{}/{}'.format(date_day, date_month, date_year)

    else:
        date_read = sys.argv[4].lower()

        if date_read == 'y':
            if date_day != 1:
                date = '{}/{}/{}'.format('0' + str(int(date_day) - 1),
                                         date_month, date_year)
            else:
                yesterday = datetime.datetime.now() - timedelta(days=1)

                date_day = yesterday.day
                date_month = yesterday.month
                if len(str(date_month)) == 1:
                    date_month = '0' + str(date_month)
                date_year = str(yesterday.year)[-2:]

                date = '{}/{}/{}'.format(date_day, date_month, date_year)

        elif date_read == 'c':
            date = input('Enter the date the book was finished (DD/MM/YY): ')

    return date


def goodreads_find():
    """Find the correct book, in the correct format, on Goodreads."""
    BROWSER.get('https://goodreads.com')

    # Login
    email_elem = BROWSER.find_element_by_name('user[email]')
    email_elem.send_keys(USERNAME)
    pass_elem = BROWSER.find_element_by_name('user[password]')
    pass_elem.send_keys(PASSWORD)
    pass_elem.send_keys(Keys.ENTER)
    time.sleep(5)

    # Find correct book and edition
    search_terms = sys.argv[1]
    search_elem = BROWSER.find_element_by_class_name('searchBox__input')
    search_elem.send_keys(search_terms + '%3Dauthor')
    search_elem.send_keys(Keys.ENTER)
    time.sleep(3)

    title_elem = BROWSER.find_element_by_class_name('bookTitle')
    title_elem.click()
    time.sleep(3)

    editions_elem = BROWSER.find_element_by_class_name('otherEditionsLink')
    editions_elem.click()

    # Format
    book_format = sys.argv[2].lower()
    filter_elem = BROWSER.find_element_by_name('filter_by_format')
    filter_elem.click()
    filter_elem.send_keys(book_format)
    filter_elem.send_keys(Keys.ENTER)
    time.sleep(3)

    book_elem = BROWSER.find_element_by_class_name('bookTitle')
    book_elem.click()

    return BROWSER.current_url


MONTH_CONV = {'01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr', '05': 'May',
              '06': 'June', '07': 'July', '08': 'Aug', '09': 'Sep', '10': 'Oct',
              '11': 'Nov', '12': 'Dec'}

def goodreads_update():
    """Update Goodreads by marking book as read and adding information."""
    # Mark as Read
    menu_elem = BROWSER.find_element_by_class_name('wtrRight.wtrUp')
    menu_elem.click()
    time.sleep(1)
    menu_elem.send_keys(Keys.TAB, Keys.TAB, Keys.TAB, Keys.ENTER)
    time.sleep(3)

    # Date Selection
    year_elem = BROWSER.find_element_by_class_name('endedAtYear.readingSession'
                                                   'DatePicker.smallPicker')
    year_elem.click()
    time.sleep(1)
    year_elem.send_keys('2', Keys.ENTER)

    month_elem = BROWSER.find_element_by_class_name('endedAtMonth.largePicker'
                                                    '.readingSessionDatePicker')
    month_elem.click()
    month_elem.send_keys(MONTH_CONV[DATE_READ[3:5]], Keys.ENTER)

    day_elem = BROWSER.find_element_by_class_name('endedAtDay.readingSession'
                                                  'DatePicker.smallPicker')
    day_elem.click()
    day_elem.send_keys(str(DATE_READ[0:2]), Keys.ENTER)

    # Save review
    save_elem = BROWSER.find_element_by_name('next')
    save_elem.click()

    # Shelf selection
    shelves_elems = BROWSER.find_elements_by_class_name('actionLinkLite.'
                                                        'bookPageGenreLink')
    shelves = []
    for shelf in shelves_elems:
        if ' users' not in shelf.text and shelf.text not in shelves:
            shelves.append(shelf.text)

    menu_elem = BROWSER.find_element_by_class_name('wtrRight.wtrDown')
    menu_elem.click()
    time.sleep(1)
    shelf_search_elem = BROWSER.find_element_by_class_name('wtrShelfSearchField')

    for i in range(len(shelves)):
        shelf_search_elem.send_keys(shelves[i], Keys.ENTER)
        shelf_search_elem.send_keys(Keys.SHIFT, Keys.HOME, Keys.DELETE)

    menu_elem.click()

    # Give star rating
    rating = sys.argv[3]
    stars_elem = BROWSER.find_elements_by_class_name('star.off')
    for stars in stars_elem:
        if stars.text.strip() == '{} of 5 stars'.format(rating):
            stars.click()
            break

    return shelves


def parse_page():
    """Parse and return page information needed for updating the spreadsheet."""
    info_list = []
    res = requests.get(URL)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    title_elem = soup.select('#bookTitle')

    rough_title = title_elem[0].getText().strip().split('\n')
    if len(rough_title) == 1:
        title = rough_title[0].strip()
    else:
        title = rough_title[0].strip() + ' ' + rough_title[2].strip()

    info_list.append(title)

    author_elem = soup.select('.authorName')
    author = author_elem[0].getText().strip()
    info_list.append(author)

    pages_elem = soup.select('#details .row')
    pages = pages_elem[0].getText().split(',')[1].strip(' pages')
    info_list.append(pages)

    if SHELVES_LIST[0] == 'Fiction' or SHELVES_LIST[0] == 'Nonfiction':
        category = SHELVES_LIST[0]
        genre = SHELVES_LIST[1]
    else:
        genre = SHELVES_LIST[0]
        if 'Fiction' in SHELVES_LIST:
            category = 'Fiction'
        else:
            category = 'Nonfiction'

    info_list.append(category)
    info_list.append(genre)

    return info_list


def input_info(sheet_name):
    """Write the book information to the first blank row on the given sheet."""
    sheet = WB.get_sheet_by_name(sheet_name)
    input_row = 1
    data = ''
    while data != None:
        data = sheet.cell(row=input_row, column=1).value
        input_row += 1

    input_row -= 1

    sheet.cell(row=input_row, column=1).value = INFO[0]        # Title
    sheet.cell(row=input_row, column=2).value = INFO[1]        # Author
    sheet.cell(row=input_row, column=3).value = int(INFO[2])   # Pages
    sheet.cell(row=input_row, column=4).value = INFO[3]        # Category
    sheet.cell(row=input_row, column=5).value = INFO[4]        # Genre
    sheet.cell(row=input_row, column=6).value = DATE_READ


CONFIG = configparser.ConfigParser()
CONFIG.read('/home/finners/Documents/Coding//Python/Booktracker/settings.ini')
USERNAME = CONFIG.get('User', 'Username')
PASSWORD = CONFIG.get('User', 'Password')
PATH = CONFIG.get('Spreadsheet', 'Path')

DATE_READ = get_date()

print('Opening a computer controlled browser window and updating Goodreads...')
BROWSER = webdriver.Firefox()

URL = goodreads_find()
SHELVES_LIST = goodreads_update()
BROWSER.close()
print('Goodreads account updated.')


WB = openpyxl.load_workbook(PATH)
print('Updating Spreadsheet...')
INFO = parse_page()

input_info('20' + DATE_READ[-2:])
input_info('Overall')

WB.save(PATH)

print('Spreadsheet has been updated.')
print('Booktracker has completed updating both the website and the spreadsheet'
      ' and will now close.')

# TODO Replace time.sleep's with Selenium waits
