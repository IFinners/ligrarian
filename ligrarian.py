#!/usr/bin/env python3

"""Automatically update Goodreads and local Spreadsheet with book read info.

Args:
    Three operational modes (g)ui, (s)earch or (u)rl

    gui arguments:
        None

    search arguments:
        Title of Book: Enclosed in double quotation marks
        Author of Book: Enclosed in double quotation marks
        Format: (p)aperback, (h)ardcover, (k)indle or (e)book
        Read Date: (t)oday, (y)esterday or a date formatted DD/MM/YY
        Rating: Number between 1 and 5
        Review (Optional): Enclosed in double quotation marks

    url arguments:
        URL: Goodreads URL for the book
        Read Date: (t)oday, (y)esterday or a date formatted DD/MM/YY
        Rating: Number between 1 and 5
        Review (Optional): Enclosed in double quotation marks
"""

import argparse
import configparser
from datetime import datetime as dt
from datetime import timedelta
import sys
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk

import bs4
import openpyxl
import requests
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait



class Gui:

    def __init__(self, master, today, yesterday):
        """Initialise GUI."""
        self.master = master
        self.today = today
        self.yesterday = yesterday

        master.title("Ligrarian")
        self.master.geometry('665x560')
        self.notebook = ttk.Notebook(master)
        self.notebook.pack(expand=1, fill="both")

        self.email = get_setting('User', 'Email')
        self.password = get_setting('User', 'Password')
        self.gui_info = {}
        self.save_choice = None

        SearchTab(self.notebook, self.master, today, yesterday)
        UrlTab(self.notebook, self.master, today, yesterday)


    def add_shared_widgets(self, tab, main_text):
        """Create widgets present on each of the notebook's tabs."""
        # Labels
        login_label = tk.Label(tab, text="Login")
        login_label.grid(row=2, column=1, sticky='W',
                         pady=(30, 5), padx=10)

        main_label = tk.Label(tab, text=main_text)
        main_label.grid(row=3, column=1, sticky='W', padx=10)

        date_label = tk.Label(tab, text="Date")
        date_label.grid(row=5, column=1, sticky='W', padx=10)

        rating_label = tk.Label(tab, text="Rating")
        rating_label.grid(row=7, column=1, sticky='W', padx=10)

        review_label = tk.Label(tab, text="Review\n (optional)",
                                     padx=10)
        review_label.grid(row=8, column=1, sticky='W')

        # Widgets
        self.login_email = tk.Entry(tab, width=20)
        if self.email:
            self.login_email.insert(0, self.email)
        else:
            self.login_email.insert(0, 'Email')
        self.login_email.grid(row=2, column=2, columnspan=3,
                              sticky='W', pady=(30, 5))

        self.login_password = tk.Entry(tab, width=20)
        if self.password:
            self.login_password.insert(0, '********')
        else:
            self.login_password.insert(0, 'Password')
        self.login_password.grid(row=2, column=4, columnspan=3, sticky='W',
                                 pady=(30, 5))

        self.save = tk.IntVar()
        self.save_box = tk.Checkbutton(tab, text='Save Password',
                                       variable=self.save, onvalue=True,
                                       offvalue=False)
        self.save_box.grid(row=2, column=7, sticky='W', pady=(30, 5))
        # Select save by default if password already saved
        if self.password:
            self.save_box.select()

        self.main = tk.Entry(tab, width=45)
        self.main.grid(row=3, column=2, columnspan=6,
                       sticky='W', pady=10)

        self.date = tk.Entry(tab, width=8)
        self.date.insert(0, self.today)
        self.date.grid(row=5, column=2, sticky='W', pady=10, ipady=3)

        self.today_button = tk.Button(tab, text="Today",
                command=lambda: self.set_date(self.date, self.today))

        self.today_button.grid(row=5, column=3, sticky='W', pady=10,)

        self.yesterday_button = tk.Button(tab, text="Yesterday",
                command=lambda: self.set_date(self.date, self.yesterday))
        self.yesterday_button.grid(row=5, column=4, sticky='W', pady=10)

        stars = ("1", "2", "3", "4", "5")
        self.star = tk.StringVar()
        self.star.set(get_setting("Defaults", "Rating"))
        self.rating = tk.OptionMenu(tab, self.star, *stars)
        self.rating.grid(row=7, column=2, sticky='W', pady=5)

        self.review = tk.Text(tab, height=15,
                              width=75, wrap=tk.WORD)
        self.review.grid(row=8, column=2, columnspan=7, sticky='W', pady=5)

        self.submit_button = tk.Button(tab, text="Mark as Read",
                                       command=self.parse_input)
        self.submit_button.grid(row=12, column=7, columnspan=2,
                                sticky='E', pady=15)

    def set_date(self, widget_name, new_value):
        """Set specified date widget to new_value."""
        widget_name.delete(0, 8)
        widget_name.insert(0, new_value)


    def parse_input(self):
        """Create input dictionary and test if required info has been given."""
        password = self.login_password.get()
        if password != '********':
            self.password = password

        self.gui_info = {
                    'email': self.login_email.get(),
                    'password': self.password,
                    'save_choice': self.save.get(),
                    'main': self.main.get(), 'date': self.date.get(),
                    'rating': self.star.get(),
                    'review': self.review.get('1.0', 'end-1c'),
        }

        if self.notebook.index("current") == 0:
            self.gui_info['format'] = self.book_format.get()

            try:
                assert self.gui_info['format']
            except AssertionError:
                messagebox.showwarning(message="Complete all non-optional "
                                       "fields before marking as read.")

        # Check shared information has been entered
        try:
            assert self.gui_info['email'] != 'Email'
            assert self.gui_info['password'] != 'Password'
            assert self.gui_info['main']
            self.master.destroy()

        except AssertionError:
            messagebox.showwarning(message="Complete all non-optional "
                                   "fields before marking as read.")

        Gui.gui_info = self.gui_info


class SearchTab(Gui):

    def __init__(self, notebook, master, today, yesterday):

        self.email = get_setting('User', 'Email')
        self.password = get_setting('User', 'Password')
        self.gui_book_info = {}
        self.save_choice = None
        self.today = today
        self.yesterday = yesterday

        self.notebook = notebook
        self.master = master

        tab = ttk.Frame(notebook)
        self.notebook.add(tab, text="Search")

        format_label = tk.Label(tab, text="Format")
        format_label.grid(row=6, column=1, sticky='W', padx=10)
        formats = ("Paperback", "Hardback", "Kindle", "Ebook",)
        self.book_format = tk.StringVar()
        self.book_format.set(get_setting("Defaults", "Format"))
        self.format = tk.OptionMenu(tab,
                                    self.book_format, *formats)
        self.format.grid(row=6, column=2, columnspan=3,
                         sticky='W', pady=5)

        self.add_shared_widgets(tab, 'Search')


class UrlTab(Gui):

    def __init__(self, notebook, master, today, yesterday):

        self.email = get_setting('User', 'Email')
        self.password = get_setting('User', 'Password')
        self.gui_book_info = {}
        self.save_choice = None
        self.today = today
        self.yesterday = yesterday

        self.notebook = notebook
        self.master = master

        self.url_tab = ttk.Frame(notebook)
        notebook.add(self.url_tab, text="URL")

        self.add_shared_widgets(self.url_tab, 'Book URL')


def parse_arguments():
    """Parse command line arguments and return dictionary of values."""
    parser = argparse.ArgumentParser(description="Goodreads updater")
    subparsers = parser.add_subparsers(help="Choose (u)rl, (s)earch or (g)ui")

    url_parser = subparsers.add_parser("url", aliases=['u'])
    url_parser.add_argument('url', metavar="url",
                            help="Book's Goodreads URL within quotes")
    url_parser.add_argument('date', help=("(t)oday, (y)esterday or "
                                          "date formatted DD/MM/YY"))
    url_parser.add_argument('rating', type=int, metavar='rating',
                            choices=[1, 2, 3, 4, 5],
                            help="A number 1 through 5")
    url_parser.add_argument('review', nargs='?', metavar="'review'",
                            help="Review enclosed in quotes")

    search_parser = subparsers.add_parser('search', aliases=['s'])
    search_parser.add_argument('terms', metavar="'terms'",
                                help="Search terms to use e.g. Book title "
                                     "and Author")
    search_parser.add_argument('format', metavar='format',
                               choices=['e', 'h', 'k', 'p'],
                               help="(p)aperback, (h)ardcover, "
                                    "(k)indle, (e)book")
    search_parser.add_argument('date', help=("(t)oday, (y)esterday or "
                                             "date formatted DD/MM/YY"))
    search_parser.add_argument('rating', type=int, metavar='rating',
                               choices=[1, 2, 3, 4, 5],
                               help="A number 1 through 5")
    search_parser.add_argument('review', nargs='?', metavar="'review'",
                               help="Review enclosed in quotes")

    gui = subparsers.add_parser("gui", aliases=['g'])
    gui.add_argument('gui', action='store_true',
                     help="Invoke GUI (Defaults to True)")

    args = parser.parse_args()
    return (vars(args))


def get_setting(section, option):
    """Return the value associated with option under section in settings"""
    config = configparser.ConfigParser()
    config.read('settings.ini')
    value = config.get(section, option)

    return value


def user_info():
    """Prompt for missing user info unless prompt is disabled."""
    email = get_setting('User', 'Email')
    if not email:
        email = input('Email: ')
    password = get_setting('User', 'Password')
    if not password:
        password = input('Password: ')
        if get_setting('Settings', 'Prompt') == 'no':
            return (email, password)

        save = input("Save Password?(y/n): ")
        if save.lower() == 'y':
            write_config(email, password, 'yes')
        elif save.lower() == 'n':
            disable = input("Disable save Password prompt?(y/n): ")
            if disable.lower() == 'y':
                write_config(email, "", 'no')
            else:
                write_config(email, "", 'yes')

    return (email, password)


def write_config(email, password, prompt):
    """Write configuration file."""
    config = configparser.ConfigParser()
    config.read('settings.ini')
    config.set('User', 'Email', email)
    config.set('User', 'Password', password)
    config.set('Settings', 'Prompt', prompt)

    with open('settings.ini', 'w') as configfile:
        config.write(configfile)


def goodreads_login(driver, email, password):
    """Login to Goodreads account from the homepage."""
    driver.get('https://goodreads.com')

    driver.find_element_by_name('user[email]').send_keys(email)
    pass_elem = driver.find_element_by_name('user[password]')
    pass_elem.send_keys(password, Keys.ENTER)

    try:
        driver.find_element_by_class_name('siteHeader__personal')
    except NoSuchElementException:
        print('Failed to login - Email and/or Password probably incorrect.')
        driver.close()
        exit()


def goodreads_find(driver, terms):
    """Find the book on Goodreads and navigate to all editions page."""
    search_elem = driver.find_element_by_class_name('searchBox__input')
    search_elem.send_keys(terms, Keys.ENTER)

    try:
        driver.find_element_by_partial_link_text('edition').click()
    except NoSuchElementException:
        print("Failed to find book using those search terms.")
        driver.close()
        exit()


def goodreads_filter(driver, book_format):
    """Filter editions with book_format and select top book."""
    pre_filter_url = driver.current_url

    # Filter by format
    filter_elem = driver.find_element_by_name('filter_by_format')
    filter_elem.click()
    filter_elem.send_keys(book_format, Keys.ENTER)

    # Make sure filtered page is loaded before clicking top book
    WebDriverWait(driver, 10).until(
        EC.url_changes((pre_filter_url))
    )

    # Select top book
    driver.find_element_by_class_name('bookTitle').click()

    return driver.current_url


def goodreads_get_shelves(driver, rating):
    """Find and return list of 'Top Shelves' on Goodreads book page."""
    shelves_elems = driver.find_elements_by_class_name('actionLinkLite.'
                                                       'bookPageGenreLink')
    shelves = []
    for shelf in shelves_elems:
        if ' users' not in shelf.text and shelf.text not in shelves:
            shelves.append(shelf.text)

    if rating == '5':
        shelves.append('5-star-books')

    return shelves


def goodreads_date_input(driver, date_done, reread):
    """Select completion date on review page, add new selectors for rereads."""
    book_code = driver.current_url.split('/')[-1]
    driver.get("https://www.goodreads.com/review/edit/{}".format(book_code))

    # If it's a reread need to create new session selectors
    if reread:
        driver.find_element_by_id('readingSessionAddLink').click()
        # More details loaded for Explicit Wait
        driver.find_element_by_class_name('smallLink.closed').click()
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, "review_recommendation"))
        )

    # Find reading session codes from all ids then use last one for date entry
    reread_codes = []
    for id_elem in driver.find_elements_by_xpath('//*[@id]'):
        id_attribute = id_elem.get_attribute('id')
        if 'readingSessionEntry' in id_attribute:
            reread_codes.append(id_attribute)
    new_read_code = reread_codes[-1][19:]

    # Date Formatting and Selection
    year = '20' + date_done[6:]
    month = date_done[3:5].lstrip('0')
    day = date_done[:2].lstrip('0')

    year_name = 'readingSessionDatePicker{}[end][year]'.format(new_read_code)
    month_name = 'readingSessionDatePicker{}[end][month]'.format(new_read_code)
    day_name = 'readingSessionDatePicker{}[end][day]'.format(new_read_code)

    Select(driver.find_element_by_name(year_name)
           ).select_by_visible_text(year)

    Select(driver.find_element_by_name(month_name)
           ).select_by_value(month)

    Select(driver.find_element_by_name(day_name)
           ).select_by_visible_text(day)


def goodreads_add_review(driver, review):
    """Write review in review box (if given)"""
    if review:
        review_elem = driver.find_element_by_name('review[review]')
        review_elem.clear()
        review_elem.click()
        review_elem.send_keys(review)


def goodreads_rate_book(driver, rating):
    """Give the book the given rating out of 5."""
    # Give star rating
    stars_elem = driver.find_elements_by_class_name('star.off')
    for stars in stars_elem:
        if stars.text.strip() == '{} of 5 stars'.format(rating):
            stars.click()
            break


def goodreads_shelve(driver, shelves):
    """Shelve book using 'Top Shelves'."""
    # Wait until review box is invisible
    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element_located((By.ID, "box"))
    )

    # Select shelves
    menu_elem = driver.find_element_by_class_name('wtrShelfButton')
    menu_elem.click()
    shelf_elem = driver.find_element_by_class_name('wtrShelfSearchField')

    for shelf in shelves:
        shelf_elem.send_keys(shelf, Keys.ENTER)
        shelf_elem.send_keys(Keys.SHIFT, Keys.HOME, Keys.DELETE)

    # Close dropdown and wait until it disappears
    menu_elem.click()
    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element_located((By.CLASS_NAME, "wtrShelfList"))
    )


def parse_page(url):
    """Parse page and return title, author and number of pages info dict."""
    info = {}
    res = requests.get(url)
    res.raise_for_status()
    soup = bs4.BeautifulSoup(res.text, 'html.parser')

    title_elem = soup.select('#bookTitle')
    rough_title = title_elem[0].getText().strip().split('\n')
    if len(rough_title) == 1:
        info['title'] = rough_title[0].strip()
    else:
        info['title'] = rough_title[0].strip() + ' ' + rough_title[2].strip()

    info['author'] = soup.select('.authorName')[0].getText().strip()

    pages_elem = soup.findAll('span', attrs={'itemprop': 'numberOfPages'})
    info['pages'] = int(pages_elem[0].getText().strip(' pages'))

    return info


def category_and_genre(shelves):
    """Determine and return category and genre from Goodreads 'Top Shelves'."""
    if 'Nonfiction' in shelves:
        category = 'Nonfiction'
    else:
        category = 'Fiction'

    for shelf in shelves:
        if shelf != category:
            genre = shelf
            break

    return (category, genre)


def input_info(year_sheet, info, date):
    """Write the book information to the first blank row on the given sheet."""
    path = get_setting('Settings', 'Path')
    workbook = openpyxl.load_workbook(path)

    existing_sheets = workbook.sheetnames
    if year_sheet not in existing_sheets:
        create_sheet(workbook, existing_sheets[-1], year_sheet)

    for sheet in [year_sheet, 'Overall']:
        sheet = workbook[sheet]

        input_row = first_blank_row(sheet)

        sheet.cell(row=input_row, column=1).value = info['title']
        sheet.cell(row=input_row, column=2).value = info['author']
        sheet.cell(row=input_row, column=3).value = info['pages']
        sheet.cell(row=input_row, column=4).value = info['category']
        sheet.cell(row=input_row, column=5).value = info['genre']
        sheet.cell(row=input_row, column=6).value = date

    workbook.save(path)


def create_sheet(workbook, last_sheet, sheet_name):
    """Create a new sheet by copying and modifying the latest one."""
    sheet = workbook.copy_worksheet(workbook[last_sheet])
    sheet.title = sheet_name
    last_row = first_blank_row(sheet)
    while last_row > 1:
        for col in range(1, 7):
            sheet.cell(row=last_row, column=col).value = None
        last_row -= 1
    day_tracker = '=(TODAY()-DATE({},1,1))/7'.format(sheet_name)
    sheet.cell(row=5, column=9).value = day_tracker


def first_blank_row(sheet):
    """Return the number of the first blank row of the given sheet."""
    input_row = 1
    data = ''
    while data is not None:
        data = sheet.cell(row=input_row, column=1).value
        input_row += 1
    input_row -= 1
    return input_row


def main():
    """Coordinate Updating of Goodreads account and writing to spreasheet."""
    today = dt.strftime(dt.now(), '%d/%m/%y')
    yesterday = dt.strftime(dt.now() - timedelta(1), '%d/%m/%y')
    args = parse_arguments()

    if 'gui' in args:
        root = tk.Tk()
        root.protocol("WM_DELETE_WINDOW", exit)
        gui = Gui(root, today, yesterday)
        root.mainloop()

        book_info = Gui.gui_info
        email = book_info['email']
        password = book_info['password']

        if book_info['save_choice']:
            write_config(email, password, 'no')
        else:
            write_config(email, "", 'yes')

    else:
        book_info = args
        email, password = user_info()
        # Process date if given as (t)oday or (y)esterday into proper format
        if book_info['date'].lower() == 't':
            book_info['date'] = today
        elif book_info['date'].lower() == 'y':
            book_info['date'] = yesterday

    print('Opening a computer controlled browser and updating Goodreads...')
    driver = webdriver.Firefox()
    driver.implicitly_wait(10)

    goodreads_login(driver, email, password)

    if 'format' in book_info:
        goodreads_find(driver, book_info['main'])
        book_url = goodreads_filter(driver, book_info['format'])
    else:
        book_url = book_info['main']
        driver.get(book_url)

    shelves = goodreads_get_shelves(driver, book_info['rating'])

    # Use rating element to see if book has been read before
    rating_elem = driver.find_element_by_class_name('stars')
    current_rating = rating_elem.get_attribute('data-rating')
    reread = current_rating != '0'

    goodreads_date_input(driver, book_info['date'], reread)
    goodreads_add_review(driver, book_info['review'])
    driver.find_element_by_name('next').click()
    driver.get(book_url)
    goodreads_rate_book(driver, book_info['rating'])
    if not reread:
        goodreads_shelve(driver, shelves)

    driver.close()
    print('Goodreads account updated.')

    print('Updating Spreadsheet...')
    info = parse_page(book_url)
    info['category'], info['genre'] = category_and_genre(shelves)
    year_sheet = '20' + book_info['date'][-2:]
    input_info(year_sheet, info, book_info['date'])

    print(('Ligrarian has completed and will now close. The following '
           'information has been written to the spreadsheet:'))
    print(info['title'], info['author'], info['pages'],
          info['category'], info['genre'], book_info['date'], sep='\n')


if __name__ == '__main__':
    main()
