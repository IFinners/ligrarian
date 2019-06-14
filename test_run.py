#!/usr/bin/env python3

"""Do a test run of the Ligrarian that resets everything afterwards."""

from datetime import datetime as dt
from datetime import timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
import time

import ligrarian


email = ligrarian.get_setting('User', 'Email')
password = ligrarian.get_setting('User', 'Password')
yesterday = dt.strftime(dt.now() - timedelta(1), '%d/%m/%y')
today = dt.strftime(dt.now(), '%d/%m/%y')
book_info = {
    'title': 'Cannery Row', 'author': 'John Steinbeck', 'date': yesterday,
    'format': 'kindle', 'rating': '4',  'review': 'Test Review',
}

driver = webdriver.Firefox()
driver.implicitly_wait(10)

ligrarian.goodreads_login(driver, email, password)
ligrarian.goodreads_find(driver, book_info['title'], book_info['author'])
url = ligrarian.goodreads_filter(driver, book_info['format'])

# Check book is correct format
info_rows = driver.find_elements_by_class_name('row')
row_text = [row.text for row in info_rows]
assert 'Kindle' in ''.join(row_text), "Book is in incorrect format."

ligrarian.goodreads_read_box(driver, book_info['date'], book_info['review'])
shelves = ligrarian.goodreads_get_shelves(driver, book_info['rating'])
ligrarian.goodreads_shelf_and_rate(driver, shelves, book_info['rating'])

time.sleep(3)

# Change date for reread
book_info['date'] = today

# Re-Read via edit URL
ligrarian.goodreads_reread(driver, book_info['date'], book_info['review'])

# Reset Goodreads account
driver.get(url)

undo_elem = driver.find_element_by_class_name('wtrStatusRead.wtrUnshelve')
assert undo_elem, "Book wasn't marked as read."
print("Book successfully marked as read.")
# Reset Goodreads account
undo_elem.click()
alert_obj = driver.switch_to.alert
alert_obj.accept()

# Check book no longer marked as read before closing window
unread_check = driver.find_element_by_class_name('wtrToRead')
print('Goodreads account reset correctly.')
driver.close()

# Spreadsheet entry testing
path = ligrarian.get_setting('Settings', 'Path')
wb = openpyxl.load_workbook(path)
print('Testing spreadsheet updating.')
info = ligrarian.parse_page(url)
info['category'], info['genre'] = ligrarian.category_and_genre(shelves)
print(info)

year_sheet = '20' + book_info['date'][-2:]

# Get first blank row before attempting to write
pre_year = ligrarian.first_blank_row(wb[year_sheet])
pre_overall = ligrarian.first_blank_row(wb['Overall'])
wb.close()

# Try to write info to both sheets
ligrarian.input_info(year_sheet, info, book_info['date'])

# Find blank rows now that the data should have been entered
wb = openpyxl.load_workbook(path)
post_year = ligrarian.first_blank_row(wb[year_sheet])
post_overall = ligrarian.first_blank_row(wb['Overall'])

# Confirm data was entered
assert pre_year < post_year, "Data not written to year sheet."
assert pre_overall < post_overall, "Data not written to overall sheet."
print("Data was written to sheet successfully.")

# Delete newly entered data
for sheet in [year_sheet, 'Overall']:
    sheet = wb[sheet]

    input_row = ligrarian.first_blank_row(sheet) - 1

    sheet.cell(row=input_row, column=1).value = ''
    sheet.cell(row=input_row, column=2).value = ''
    sheet.cell(row=input_row, column=3).value = ''
    sheet.cell(row=input_row, column=4).value = ''
    sheet.cell(row=input_row, column=5).value = ''
    sheet.cell(row=input_row, column=6).value = ''

wb.save(path)

print("Spreadsheet reset.")
print("Test run complete.")
